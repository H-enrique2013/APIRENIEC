from PySide2.QtWidgets import QTableWidget,QTableWidgetItem,QHeaderView,QAbstractItemView,QPushButton, QWidget, QFileDialog,QCompleter
from PySide2.QtCore import Qt, QDate
from PySide2.QtGui import *
from controllers.AlignDelegate import *
from views.new_padron import Ui_NewBook
#from db.books import EncontrarDni,UpdateAutoriza
from PySide2 import QtCore
from pys2_msg_boxes import msg_boxes
from datetime import date, datetime
import math
import shutil
import os
import win32com.client as win32
from openpyxl import load_workbook
import pandas as pd



class NewBookWindow(QWidget,Ui_NewBook):
    def __init__(self, parent=None,_dataframePadron=None):
        super().__init__(parent)
        self.parent = parent
        self.dataframePadron = _dataframePadron
        self.setupUi(self)
        self.setWindowFlag(Qt.Window)
        #self.setMaximumSize(QtCore.QSize(1212, 562))
        #self.populate_comboboxAreaVisitada()
        self.inputSuperficie.setEnabled(False)
        self.inputMonto.setEnabled(False)

        '''
        self.onlyInt = QIntValidator()
        self.dniLineEdit.setValidator(self.onlyInt)

        '''
        print("DATAFRAME PADRON:")
        print(self.dataframePadron)
        
        self.btnExcel.clicked.connect(self.generar_padron_excel)
        self.addPdf.clicked.connect(self.generar_padron_pdf)
        self.btnGenerarPadron.clicked.connect(self.generar_padron)
        self.cancelButton.clicked.connect(self.close)

        
        self.tableMasiva2.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tableMasiva2.verticalHeader().setVisible(False)
        self.tableMasiva2.setColumnWidth(0,19)


        headerVertical4 = self.tableMasiva2.verticalHeader()
        headerVertical4.resizeSections(QHeaderView.ResizeToContents)
        headerVertical4.setStretchLastSection(True)

        delegate4= AlignDelegate(self.tableMasiva2)

        self.tableMasiva2.setItemDelegateForColumn(0,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(1,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(2,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(4,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(5,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(6,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(7,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(8,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(9,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(10,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(11,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(12,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(13,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(14,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(15,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(16,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(17,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(18,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(19,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(20,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(21,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(22,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(23,delegate4)
        self.tableMasiva2.setItemDelegateForColumn(24,delegate4)
        self.tableMasiva2.cellChanged.connect(self.cell_was_edited)

        self.table_config_masivo2()

    def table_config_masivo2(self):
        
        #column_headers = ("N°","DNI", "AP_PAT", "AP_MAT","NOMBRES", "GENERO","FECHA_NAC", "DIRECCION","UBIGEO_NAC","UBIGEO_DIR", "PADRE", "MADRE", "EST_CIVIL")
        column_headers = ("N°", "N° DNI", "APELLIDO\nPATERNO", "APELLIDO\nMATERNO", "NOMBRES", "SEXO","FECHA DE\nNACIMIENTO", "UBIGEO", "DOMICILIO DE\nBENEFICIARIO", "DISTRITO","PROVINCIA", "SUPERFICIE\nINDEMNIZABLE\n(ha)", "MONTO\nINDEMNIZABLE\n(S/)")
        self.tableMasiva2.setColumnCount(len(column_headers))
        self.tableMasiva2.setHorizontalHeaderLabels(column_headers)
        
        stylesheet = "::section{Background-color:red;border:1px solid black;color: white;}"
        self.tableMasiva2.horizontalHeader().setStyleSheet(stylesheet)

        self.tableMasiva2.setSelectionBehavior(QAbstractItemView.SelectItems)
        self.tableMasiva2.setStyleSheet("QTableWidget{ gridline-color:black;alternate-background-color: #E2DED0;}")
        self.tableMasiva2.setAlternatingRowColors(True)
        self.tableMasiva2.setEditTriggers(QTableWidget.DoubleClicked)
    
    def cell_was_edited(self, row, column):
        item = self.tableMasiva2.item(row, column)
        if item is not None:
            new_value = item.text()
            print(f'Cell at row {row}, column {column} was edited. New value: {new_value}')
            # Aquí puedes agregar lógica adicional para manejar el nuevo valor si es necesario

    def generar_padron(self):

        dataframeDepurado = self.dataframePadron
        inputSuperficie = dataframeDepurado['Superficie'].sum()
        inputMonto = dataframeDepurado['Monto_Indemnizable'].sum()
        self.inputSuperficie.setText(str(inputSuperficie))
        self.inputMonto.setText(str(inputMonto))
        
        
        dataframeDepurado = dataframeDepurado.rename(columns={
            'DNI': 'N° DNI',
            'AP_PAT': 'APELLIDO PATERNO',
            'AP_MAT': 'APELLIDO MATERNO',
            'FECHA_NAC': 'FECHA DE NACIMIENTO',
            'UBIGEO_DIR': 'UBIGEO',
            'DIRECCION': 'DOMICILIO DE BENEFICIARIO',
            'DISTRITO_D': 'DISTRITO',
            'PROVINCIA_D': 'PROVINCIA',
            'Superficie': 'SUPERFICIE INDEMNIZABLE (ha)',
            'Monto_Indemnizable':'MONTO INDEMNIZABLE (S/)'
        })

        #dataframeDepurado["N°"] = range(1, len(dataframeDepurado) + 1)
        # Ordenar el DataframeDepurado
        dataframeDepurado = dataframeDepurado[['N° DNI', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES', 'SEXO',
                                            'FECHA DE NACIMIENTO', 'UBIGEO', 'DOMICILIO DE BENEFICIARIO', 'DISTRITO',
                                            'PROVINCIA', 'SUPERFICIE INDEMNIZABLE (ha)', 'MONTO INDEMNIZABLE (S/)']]
        
        DataTuplas = dataframeDepurado.values.tolist()
      
        # Cambios últimos
        self.populate_table_masivo2(DataTuplas)


    #def generar_padron_excel(self):
#
    #    row_count = self.tableMasiva2.rowCount()
    #    column_count = self.tableMasiva2.columnCount()
    #    
    #    data = []
    #    for row in range(row_count):
    #        row_data = []
    #        for column in range(column_count):
    #            item = self.tableMasiva2.item(row, column)
    #            row_data.append(item.text() if item is not None else "")
    #        data.append(row_data)
    #    
    #    column_headers = [self.tableMasiva2.horizontalHeaderItem(i).text() for i in range(column_count)]
    #    df = pd.DataFrame(data, columns=column_headers)
#
    #    dataframeDepurado = df
    #    inputRegion = self.inputRegion.text()
    #    inputProvincia = self.inputProvincia.text()
    #    inputDistrito = self.inputDistrito.text()
    #    inputSector = self.inputSector.text()
    #    inputCultivo = self.inputCultivo.text()
    #    inputSuperficie = dataframeDepurado['SUPERFICIE INDEMNIZABLE (ha)'].sum()
    #    inputMonto = dataframeDepurado['MONTO INDEMNIZABLE (S/)'].sum()
#
#
    #    #dataframeDepurado = dataframeDepurado.rename(columns={
    #    #    'DNI': 'N° DNI',
    #    #    'AP_PAT': 'APELLIDO PATERNO',
    #    #    'AP_MAT': 'APELLIDO MATERNO',
    #    #    'FECHA_NAC': 'FECHA DE NACIMIENTO',
    #    #    'UBIGEO_DIR': 'UBIGEO',
    #    #    'DIRECCION': 'DOMICILIO DE BENEFICIARIO',
    #    #    'DISTRITO_D': 'DISTRITO',
    #    #    'PROVINCIA_D': 'PROVINCIA',
    #    #    'Superficie': 'Superficie Indemnizable (ha)',
    #    #    'Monto_Indemnizable':'MONTO INDEMNIZABLE (S/)'
    #    #})
#
    #    #dataframeDepurado["N°"] = range(1, len(dataframeDepurado) + 1)
    #    
    #    #dataframeDepurado = dataframeDepurado[['N°', 'N° DNI', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES', 'SEXO',
    #    #                                    'FECHA DE NACIMIENTO', 'UBIGEO', 'DOMICILIO DE BENEFICIARIO', 'DISTRITO',
    #    #                                    'PROVINCIA', 'Superficie Indemnizable (ha)', 'MONTO INDEMNIZABLE (S/)']]
    #    
#
    #    now = datetime.now()
    #    # dd/mm/YY H:M:S
    #    dt_string = now.strftime("%d/%m/%Y %H:%M:%S")
    #    dt_string=dt_string.replace(':','_')
    #    dt_string=dt_string.replace('/','-')
    #    dt_string_final = inputRegion +'_'+inputSector +'_'+ inputCultivo + '_'+dt_string
    #    nombre_archivo = f'./REPORTES/{dt_string_final}.xlsx'
#
    #    columnas = [
    #        'N°', 'N° DNI', 'APELLIDO PATERNO', 'APELLIDO MATERNO', 'NOMBRES', 'SEXO', 
    #        'FECHA DE NACIMIENTO', 'UBIGEO', 'DOMICILIO DE BENEFICIARIO', 'DISTRITO', 
    #        'PROVINCIA', 'Superficie Indemnizable (ha)', 'Monto Indemnizable (S/)'
    #    ]
#
    #    with pd.ExcelWriter(nombre_archivo, engine='xlsxwriter') as writer:
    #        workbook = writer.book
    #        worksheet = workbook.add_worksheet('Padrón')
#
    #        # Agregar casillas adicionales con datos de entrada
    #        additional_data = [
    #            ('REGIÓN', inputRegion),
    #            ('PROVINCIA', inputProvincia),
    #            ('DISTRITO', inputDistrito),
    #            ('SECTOR ESTADISTICO', inputSector),
    #            ('CULTIVO INDEMNIZABLE', inputCultivo),
    #            ('SUPERFICIE INDEMNIZABLE (ha)', inputSuperficie),
    #            ('MONTO INDEMNIZABLE (s/)', inputMonto)
    #        ]
    #        
    #        row = 0
    #        for key, value in additional_data:
    #            worksheet.write(row, 0, key)
    #            worksheet.write(row, 1, value)
    #            row += 1
#
    #        # Agregar espacio entre datos adicionales y encabezados del dataframe
    #        row += 1
#
    #        # Escribir encabezados del DataFrame
    #        formato_encabezado = workbook.add_format({
    #            'bold': True, 'border': 1, 'font_color': 'white', 'font_name': 'Cambria', 'font_size': 10,
    #            'bg_color': '#4F81BD', 'align': 'center', 'valign': 'vcenter', 'text_wrap': True
    #        })
#
    #        for col_num, encabezado in enumerate(columnas):
    #            worksheet.write(row, col_num, encabezado, formato_encabezado)
#
    #        # Escribir datos del DataFrame
    #        formato_general = workbook.add_format({
    #            'font_name': 'Cambria', 'border': 1, 'font_size': 10, 'align': 'center', 'valign': 'vcenter', 'text_wrap': True
    #        })
#
    #        for i, data_row in dataframeDepurado.iterrows():
    #            for j, value in enumerate(data_row):
    #                value = self.limpiar_dato(value)
    #                worksheet.write(row + i + 1, j, value, formato_general)
#
    #        anchos_columnas = [12.43, 15, 20, 20, 20, 10, 15, 10, 25, 15, 15, 25, 20]
    #        for col_num, ancho in enumerate(anchos_columnas):
    #            worksheet.set_column(col_num, col_num, ancho)
#
    #        worksheet.set_zoom(80)
    #    
    #    print(f"Archivo Excel guardado como {nombre_archivo}")
    #    msg_boxes.correct_msgbox("¡Generación Exitosa!","¡Se ha guardado el archivo en la carpeta REPORTES!")
#
#
       #    CONTINUA EL CODIGO
    
    def generar_padron_excel(self):
        row_count = self.tableMasiva2.rowCount()
        column_count = self.tableMasiva2.columnCount()
        print("CANTIDAD DE COLUMNAS: " + str(column_count))

        inputRegion = self.inputRegion.text()
        inputProvincia = self.inputProvincia.text()
        inputDistrito = self.inputDistrito.text()
        inputSector = self.inputSector.text()
        inputCultivo = self.inputCultivo.text()
        inputFechaTermino = self.inputFechaTermino.text()
        inputSuperficie = self.inputSuperficie.text()
        inputMonto = self.inputMonto.text()
        UbigeoEncontrado=self.BuscarUbigeo(inputDistrito,inputProvincia,inputRegion)

        data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = self.tableMasiva2.item(row, column)
                print("EL COLUM ES... : " + str(column))
                if item is not None:
                    if column == 7:  # Columna H (7)
                        row_data.append(UbigeoEncontrado) if item.text() in ["", " "] else row_data.append(item.text())
                    elif column == 8:  # Columna I (8)
                        row_data.append("S/D") if item.text() in ["", " "] else row_data.append(item.text())
                    elif column == 9 and item.text() in ["", " "]:  # Columna J (9)
                        row_data.append(str(inputDistrito))
                    elif column == 10 and item.text() in ["", " "]:  # Columna K (10)
                        row_data.append(str(inputProvincia))
                    else:
                        row_data.append(item.text())
                else:
                    row_data.append("")
            print(f"Row {row} data length: {len(row_data)}")  # Debugging line
            data.append(row_data)

        # Check the length of each row to ensure they all match column_count
        for i, row_data in enumerate(data):
            if len(row_data) != column_count:
                print(f"Row {i} length mismatch: expected {column_count}, got {len(row_data)}")
                return  # Exit the function if there's a mismatch

        print("CANTIDAD DE COLUMNAS: " + str(column_count))
        column_headers = [self.tableMasiva2.horizontalHeaderItem(i).text() for i in range(column_count)]
        print("HEADERS: " + str(column_headers))

        df = pd.DataFrame(data, columns=column_headers)

        dataframeDepurado = df

        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y_%H_%M_%S")
        dt_string_final = f"{inputRegion}_{inputSector}_{inputCultivo}_{dt_string}"
        suggested_filename = f"{dt_string_final}.xlsx"

        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(self, "Guardar archivo", suggested_filename, "Archivos Excel (*.xlsx);;Todos los archivos (*)", options=options)

        if not file_path:
            return

        template_path = './assets/PlantillaFormato01.xlsx'
        workbook = load_workbook(template_path)
        worksheet = workbook.active

        worksheet['D3'] = inputRegion
        worksheet['D4'] = inputProvincia
        worksheet['D5'] = inputDistrito
        worksheet['D6'] = inputSector
        worksheet['I4'] = inputCultivo
        worksheet['I5'] = inputSuperficie
        worksheet['I6'] = inputMonto
        worksheet['A3498'] = "Fecha de termino padron: " + str(inputFechaTermino)

        start_row = 8
        filafinal = 0
        for i, data_row in dataframeDepurado.iterrows():
            for j, value in enumerate(data_row):
                filafinal = start_row + i
                worksheet.cell(row=start_row + i, column=j + 1, value=value)

        last_data_row = filafinal + 1
        print("FILA FINAL:")
        print(str(last_data_row))
        if last_data_row <= 3494:
            for row in range(last_data_row, 3496):
                worksheet.row_dimensions[row].hidden = True

        workbook.save(file_path)
        if file_path:
            msg_boxes.correct_msgbox("¡Generación Exitosa!", f"¡Se ha guardado el archivo en {file_path}!")
        else:
            msg_boxes.error_msgbox("¡Guardado Cancelado!", "El archivo PDF no se ha guardado.")


    def generar_padron_pdf(self):
        row_count = self.tableMasiva2.rowCount()
        column_count = self.tableMasiva2.columnCount()

        inputRegion = self.inputRegion.text()
        inputProvincia = self.inputProvincia.text()
        inputDistrito = self.inputDistrito.text()
        inputFechaTermino = self.inputFechaTermino.text()
        inputSector = self.inputSector.text()
        inputCultivo = self.inputCultivo.text()
        inputSuperficie = self.inputSuperficie.text()
        inputMonto = self.inputMonto.text()
        UbigeoEncontrado=self.BuscarUbigeo(inputDistrito,inputProvincia,inputRegion)
        
        data = []
        for row in range(row_count):
            row_data = []
            for column in range(column_count):
                item = self.tableMasiva2.item(row, column)
                print("EL COLUM ES... : " + str(column))
                if item is not None:
                    if column == 7:  # Columna H (7)
                        row_data.append(UbigeoEncontrado) if item.text() in ["", " "] else row_data.append(item.text())
                    elif column == 8:  # Columna I (8)
                        row_data.append("S/D") if item.text() in ["", " "] else row_data.append(item.text())
                    elif column == 9 and item.text() in ["", " "]:  # Columna J (9)
                        row_data.append(str(inputDistrito))
                    elif column == 10 and item.text() in ["", " "]:  # Columna K (10)
                        row_data.append(str(inputProvincia))
                    else:
                        row_data.append(item.text())
                else:
                    row_data.append("")
            print(f"Row {row} data length: {len(row_data)}")  # Debugging line
            data.append(row_data)

        # Check the length of each row to ensure they all match column_count
        for i, row_data in enumerate(data):
            if len(row_data) != column_count:
                print(f"Row {i} length mismatch: expected {column_count}, got {len(row_data)}")
                return  # Exit the function if there's a mismatch

        print("CANTIDAD DE COLUMNAS: " + str(column_count))
        column_headers = [self.tableMasiva2.horizontalHeaderItem(i).text() for i in range(column_count)]
        print("HEADERS: " + str(column_headers))

        df = pd.DataFrame(data, columns=column_headers)

        dataframeDepurado = df
        

        now = datetime.now()
        dt_string = now.strftime("%d-%m-%Y_%H_%M_%S")
        dt_string_final = f"{inputRegion}_{inputSector}_{inputCultivo}_{dt_string}"
        excel_temp_path = './assets/PadronTemporal.xlsx'
        pdf_output_name = dt_string_final + '.pdf'

        # Cargar plantilla
        template_path = './assets/PlantillaFormato01.xlsx'
        workbook = load_workbook(template_path)
        worksheet = workbook['Hoja1']

        # Escribir valores en las celdas especificadas
        worksheet['D3'] = inputRegion
        worksheet['D4'] = inputProvincia
        worksheet['D5'] = inputDistrito
        worksheet['D6'] = inputSector
        worksheet['I4'] = inputCultivo
        worksheet['I5'] = inputSuperficie
        worksheet['I6'] = inputMonto
        worksheet['A3498'] = "Fecha de termino padron: " + str(inputFechaTermino)
        # Escribir el DataFrame a partir de A8 sin encabezados
        start_row = 8
        filafinal = 0
        for i, data_row in dataframeDepurado.iterrows():
            for j, value in enumerate(data_row):
                filafinal = start_row + i
                worksheet.cell(row=start_row + i, column=j + 1, value=value)

        # Ocultar filas vacías hasta A36 si no contienen datos
        last_data_row = filafinal + 1
        print("FILA FINAL:")
        print(str(last_data_row))
        if last_data_row <= 3494:
            for row in range(last_data_row, 3496):
                worksheet.row_dimensions[row].hidden = True

        # Guardar el archivo Excel temporal
        try:
            workbook.save(excel_temp_path)
        except Exception as e:
            print(f"Error al guardar el archivo Excel: {e}")
            msg_boxes.error_msgbox("Error", "Se produjo un error al guardar el archivo Excel.")
            return

        # Convertir Excel a PDF
        excel = win32.Dispatch('Excel.Application')
        excel.Visible = False
        try:
            wb = excel.Workbooks.Open(os.path.abspath(excel_temp_path))
        except Exception as e:
            print(f"Error al abrir el archivo Excel con win32com: {e}")
            msg_boxes.error_msgbox("Error", "Se produjo un error al abrir el archivo Excel con win32com.")
            excel.Quit()
            return

        if not wb:
            print("Error: No se pudo abrir el archivo Excel.")
            msg_boxes.error_msgbox("Error", "No se pudo abrir el archivo Excel.")
            excel.Quit()
            return

        ws = wb.Sheets['Hoja1']

        # Configurar página en orientación horizontal, tamaño A4 y ajustar todas las columnas a una página
        try:
            ws.PageSetup.Orientation = 2  # xlLandscape
            ws.PageSetup.PaperSize = 9    # xlPaperA4
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1
            ws.PageSetup.FitToPagesTall = False

            pdf_temp_path = os.path.abspath('./assets/PadronTemporal.pdf')
            wb.ExportAsFixedFormat(0, pdf_temp_path)
            wb.Close(SaveChanges=False)
            excel.Quit()
        except Exception as e:
            print(f"Error al exportar el archivo a PDF: {e}")
            msg_boxes.error_msgbox("Error", "Se produjo un error al exportar el archivo a PDF.")
            wb.Close(SaveChanges=False)
            excel.Quit()
            return

        print(f"Archivo PDF guardado como {pdf_temp_path}")

        # Preguntar al usuario dónde guardar el PDF
        options = QFileDialog.Options()
        file_path, _ = QFileDialog.getSaveFileName(None, "Guardar PDF", pdf_output_name, "PDF Files (*.pdf)", options=options)

        if file_path:
            shutil.move(pdf_temp_path, file_path)  # Usar shutil.move en lugar de os.rename
            msg_boxes.correct_msgbox("¡Generación Exitosa!", f"¡Se ha guardado el archivo en {file_path}!")
        else:
            os.remove(pdf_temp_path)
            msg_boxes.error_msgbox("¡Guardado Cancelado!", "El archivo PDF no se ha guardado.")
                    
    def clean_inputs(self):
        self.inputRegion.clear()
        self.inputProvincia.clear()
        self.inputDistrito.clear()
        self.inputSector.clear()
        self.inputCultivo.clear()
        self.inputSuperficie.clear()
        self.inputMonto.clear()
       
    def limpiar_dato(self,valor):
        if isinstance(valor, float) and (math.isnan(valor) or math.isinf(valor)):
            return ""
        return valor
    
    def populate_table_masivo2(self,data):
        self.tableMasiva2.setRowCount(len(data))
        print("REFRESH4")
        for index_row, row in enumerate(data):

            # Incrementamos el índice de fila en 1 para comenzar con el número de orden en 1
            order_number = index_row + 1
            
            # Establecemos el número de orden en la primera columna
            self.tableMasiva2.setItem(index_row, 0, QTableWidgetItem(str(order_number)))
            
            for index_cell, cell in enumerate(row):
                if cell is None or cell == "NULL" or (isinstance(cell, float) and math.isnan(cell)):
                    cell = ""
                # Llenamos los datos de la fila a partir de la segunda columna
                if index_cell == 4: 
                    print("LA CELDA ES:") 
                    print(cell) 
                    if cell == "1":
                        cell_text = "H"
                    elif cell == "2":
                        cell_text = "M"
                    else:
                        cell_text = str(cell)  # Mantener el valor original si no es 1 ni 2
                    self.tableMasiva2.setItem(index_row, index_cell + 1, QTableWidgetItem(cell_text))
                else:
                    self.tableMasiva2.setItem(index_row, index_cell + 1, QTableWidgetItem(str(cell)))
        print("REFRESH5")
        headerVertical = self.tableMasiva2.verticalHeader()
        headerVertical.resizeSections(QHeaderView.ResizeToContents)
        headerVertical.setStretchLastSection(True)


    def BuscarUbigeo(self,distrito,provincia,departamento):
        #Cargar Ubigeos
        # Cargar el archivo Excel
        df = pd.read_excel("geodir-ubigeo-reniec.xlsx")
        
        # Filtrar los datos según los parámetros dados
        filtro = (df['Distrito'] == distrito) & (df['Provincia'] == provincia) & (df['Departamento'] == departamento)
        resultado = df.loc[filtro, 'Ubigeo']
        
        # Verificar si se encontraron resultados
        if len(resultado) > 0:
            return resultado.iloc[0]  # Devolver el primer valor encontrado
        else:
            return "No se encontró el ubigeo para los parámetros dados"

        